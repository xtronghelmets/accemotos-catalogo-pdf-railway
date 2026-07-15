# -*- coding: utf-8 -*-
"""
lector_maestro.py

Lee data/data.xlsx (hoja "Hoja Catalogo"), la fuente de verdad de inventario,
precios, marca, categoría de catálogo, talla e imagen, y produce, para cada
catálogo definido en categorias_config.CATALOGOS, la lista de grupos
(Grupo_Foto) que deben aparecer, ya ordenada por subcategoría.

Esquema de la hoja "Hoja Catalogo":
    A  SKU
    B  Grupo_Foto
    C  Precio mayor
    D  Precio Detal
    E  INVENTARIO
    F  MARCA
    G  Nombre_catalogo
    H  talla
    I  categoria catalogo   → define el catálogo/botón (ver categorias_config)
    J  subategoria Catalogo  → sección/encabezado dentro del PDF
    K  IMAGEN 1              → URL(s) de imagen separadas por coma

Reglas aplicadas:
- Solo entran grupos con AL MENOS un SKU con inventario > 0.
- Solo entran grupos que tienen AL MENOS una URL de imagen (columna K). Los
  productos sin URL no se incluyen en el catálogo (decisión del negocio).
- El catálogo de cada grupo lo decide categorias_config.catalogo_asignado()
  a partir de la marca (columna F) y la "categoria catalogo" (columna I).
- La sección dentro del PDF es la "subategoria Catalogo" (columna J).
- La talla sale de la columna H tal cual ('Sin talla' se trata como talla única).
"""
import os
import openpyxl

from categorias_config import (
    normalizar_marca, catalogo_asignado, orden_subcategoria, CATALOGOS,
)

HOJA_PREFERIDA = 'Hoja Catalogo'

# Rutas donde puede vivir el Excel, en orden de preferencia.
def _rutas_candidatas(base_dir):
    return [
        os.path.join(base_dir, 'data', 'data.xlsx'),
        os.path.join(base_dir, '..', 'data', 'data.xlsx'),
        os.path.join('/var/task', 'data', 'data.xlsx'),
        os.path.join(base_dir, 'data.xlsx'),  # fallback raíz
    ]


def _normalizar_talla(valor):
    """Columna H → talla mostrable. 'Sin talla'/vacío → '' (talla única)."""
    if valor is None:
        return ''
    t = str(valor).strip()
    if not t or t.lower() in ('sin talla', 'nan', 'none', '-'):
        return ''
    return t.upper()


def _tiene_url(valor):
    if not valor:
        return False
    s = str(valor).strip()
    return s.lower() not in ('', 'nan', 'none') and 'http' in s.lower()


def _encontrar_excel(base_dir, ruta_excel_forzada=None, log=lambda m: None):
    if ruta_excel_forzada and os.path.exists(ruta_excel_forzada):
        return ruta_excel_forzada
    for cand in _rutas_candidatas(base_dir):
        if os.path.exists(cand):
            log(f"  📊 Excel de catálogo: {cand}")
            return cand
    return None


def _indices_columnas(headers):
    """Mapea nombres de encabezado (flexibles) a índices de columna."""
    def buscar(*alias):
        for i, h in enumerate(headers):
            hl = str(h).strip().lower()
            if hl in [a.lower() for a in alias]:
                return i
        return -1

    return {
        'sku':          buscar('SKU'),
        'grupo':        buscar('Grupo_Foto', 'grupo_foto', 'grupo foto'),
        'pmayor':       buscar('Precio mayor', 'Precio_Mayor', 'precio mayor'),
        'pdetal':       buscar('Precio Detal', 'Precio_Detal', 'precio detal'),
        'inv':          buscar('INVENTARIO', 'inventario', 'stock'),
        'marca':        buscar('MARCA', 'marca'),
        'nombre':       buscar('Nombre_catalogo', 'nombre_catalogo', 'nombre'),
        'talla':        buscar('talla', 'Talla'),
        'cat_catalogo': buscar('categoria catalogo', 'categoría catalogo',
                               'categoria_catalogo'),
        'subcategoria': buscar('subategoria Catalogo', 'subcategoria catalogo',
                               'subcategoría catalogo', 'subategoria catalogo'),
        'imagen':       buscar('IMAGEN 1', 'imagen 1', 'imagen', 'imagenes',
                               'imágenes'),
    }


def cargar_grupos_activos(base_dir, ruta_excel=None, callback_log=None):
    """
    Lee el Excel completo y devuelve:
    {
      'grupos': { grupo_foto: {...} },   # solo grupos activos y válidos
      'ruta_excel': ruta,
      'omitidos': {
          'sin_grupo_foto': int,
          'sin_inventario': int,
          'sin_imagen': int,
          'catalogo_no_reconocido': set([...]),
      }
    }
    """
    def log(m):
        if callback_log:
            callback_log(m)

    ruta = _encontrar_excel(base_dir, ruta_excel, log)
    if not ruta:
        raise FileNotFoundError(
            "No se encontró data/data.xlsx. "
            "Verifica que el archivo esté en la carpeta data/ del repo."
        )

    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb[HOJA_PREFERIDA] if HOJA_PREFERIDA in wb.sheetnames else wb.worksheets[0]
    filas = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h is not None else '' for h in filas[0]]
    col = _indices_columnas(headers)

    def val(fila, key):
        i = col[key]
        return fila[i] if (i is not None and i >= 0 and i < len(fila)) else None

    grupos = {}
    omitidos = {
        'sin_grupo_foto': 0,
        'sin_inventario': 0,
        'sin_imagen': 0,
        'catalogo_no_reconocido': set(),
    }

    for fila in filas[1:]:
        if not fila or all(v is None for v in fila):
            continue

        grupo_foto = val(fila, 'grupo')
        if not grupo_foto:
            omitidos['sin_grupo_foto'] += 1
            continue
        grupo_foto = str(grupo_foto).strip()

        sku          = val(fila, 'sku')
        nombre       = (str(val(fila, 'nombre') or '')).strip()
        talla        = _normalizar_talla(val(fila, 'talla'))
        cat_catalogo = (str(val(fila, 'cat_catalogo') or '')).strip()
        subcategoria = (str(val(fila, 'subcategoria') or '')).strip()
        imagen       = (str(val(fila, 'imagen') or '')).strip()
        if imagen.lower() in ('nan', 'none'):
            imagen = ''

        inv_raw = val(fila, 'inv')
        inventario = int(inv_raw) if isinstance(inv_raw, (int, float)) else 0

        precio_mayor = val(fila, 'pmayor')
        precio_detal = val(fila, 'pdetal')
        marca = normalizar_marca(val(fila, 'marca'))

        catalogo_id = catalogo_asignado(marca, cat_catalogo)
        if catalogo_id not in CATALOGOS:
            if cat_catalogo:
                omitidos['catalogo_no_reconocido'].add(cat_catalogo)
            continue

        if grupo_foto not in grupos:
            grupos[grupo_foto] = {
                'grupo_foto':      grupo_foto,
                'marca':           marca or CATALOGOS[catalogo_id]['marca'],
                'catalogo':        catalogo_id,
                'subcategoria':    subcategoria,
                'nombre_producto': nombre,
                'imagen_url':      '',
                'skus':            [],
            }
        g = grupos[grupo_foto]

        # Primera imagen no vacía encontrada para el grupo.
        if not g['imagen_url'] and _tiene_url(imagen):
            g['imagen_url'] = imagen
        # Completar nombre/subcategoría si faltaban.
        if not g['nombre_producto'] and nombre:
            g['nombre_producto'] = nombre
        if not g['subcategoria'] and subcategoria:
            g['subcategoria'] = subcategoria

        g['skus'].append({
            'sku':          sku,
            'nombre_producto': nombre,
            'talla':        talla,
            'inventario':   inventario,
            'precio_mayor': precio_mayor,
            'precio_detal': precio_detal,
            'imagen_url':   imagen if _tiene_url(imagen) else '',
        })

    # ── Filtrado final ───────────────────────────────────────────────────────
    grupos_activos = {}
    for gf, g in grupos.items():
        if not any(s['inventario'] > 0 for s in g['skus']):
            omitidos['sin_inventario'] += 1
            continue
        if not _tiene_url(g['imagen_url']):
            omitidos['sin_imagen'] += 1
            log(f"  🚫 Excluido (sin URL de imagen): {gf} — "
                f"{g['nombre_producto'][:40]}")
            continue
        grupos_activos[gf] = g

    log(f"  ✅ {len(grupos_activos)} grupos activos de {len(grupos)} grupos totales")
    if omitidos['sin_imagen']:
        log(f"  🚫 {omitidos['sin_imagen']} grupos excluidos por no tener URL de imagen")
    if omitidos['catalogo_no_reconocido']:
        log(f"  ⚠️ Categorías de catálogo (columna I) no reconocidas: "
            f"{sorted(omitidos['catalogo_no_reconocido'])}")

    return {'grupos': grupos_activos, 'ruta_excel': ruta, 'omitidos': omitidos}


def _nombre_para_ordenar(grupo):
    """Nombre 'limpio' usado para ordenar alfabéticamente dentro de la sección."""
    import re
    nombre = grupo['nombre_producto'] or ''
    talla = grupo['skus'][0]['talla'] if grupo['skus'] else ''
    if talla:
        nombre = re.sub(re.escape(talla), '', nombre, flags=re.IGNORECASE)
    return re.sub(r'\s+', ' ', nombre).strip().upper()


def armar_catalogo(catalogo_id, grupos_activos, callback_log=None):
    """
    Devuelve una lista ordenada de secciones para un catálogo específico:
    [(subcategoria_key, subcategoria_label, [grupo, grupo, ...]), ...]

    - Filtra los grupos cuyo 'catalogo' asignado coincide con catalogo_id.
    - Agrupa por subcategoría (columna J) y ordena las secciones según
      categorias_config.orden_subcategoria(); dentro de cada sección, por nombre.
    """
    def log(m):
        if callback_log:
            callback_log(m)

    cfg = CATALOGOS.get(catalogo_id)
    if not cfg:
        raise ValueError(f"Catálogo desconocido: {catalogo_id}")

    del_catalogo = [g for g in grupos_activos.values()
                    if g.get('catalogo') == catalogo_id]

    # Agrupar por subcategoría (columna J)
    secciones = {}
    for g in del_catalogo:
        clave = g.get('subcategoria') or 'Otros'
        secciones.setdefault(clave, []).append(g)

    resultado = []
    for subcat in sorted(secciones.keys(), key=orden_subcategoria):
        grupos = sorted(secciones[subcat], key=_nombre_para_ordenar)
        resultado.append((subcat, subcat, grupos))

    total_grupos = sum(len(x[2]) for x in resultado)
    log(f"  📚 Catálogo '{cfg['nombre']}': {total_grupos} grupos en "
        f"{len(resultado)} secciones")
    return resultado
